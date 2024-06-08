import openpyxl
import os
from openpyxl import load_workbook
from copy import copy

# Specify the path to your original Excel file
original_file_path = 'xls/Chelsea Ng_Java-B.xlsx'

# Create the "report card" directory if it doesn't exist
output_directory = 'report card'
if not os.path.exists(output_directory):
    os.makedirs(output_directory)

# Load the original workbook
original_workbook = load_workbook(original_file_path)

# Iterate over each sheet and save it as a new Excel file in the "report card" folder
for sheet_name in original_workbook.sheetnames:
    # Create a new workbook
    new_workbook = openpyxl.Workbook()
    new_workbook.remove(new_workbook.active)  # Remove the default sheet created with new workbook

    # Get the original sheet
    original_sheet = original_workbook[sheet_name]

    # Create a new sheet in the new workbook
    new_sheet = new_workbook.create_sheet(title=sheet_name)

    # Copy dimensions (row heights and column widths)
    for col_dim, col_dim_value in original_sheet.column_dimensions.items():
        new_sheet.column_dimensions[col_dim] = copy(col_dim_value)
    for row_dim, row_dim_value in original_sheet.row_dimensions.items():
        new_sheet.row_dimensions[row_dim] = copy(row_dim_value)

    # Copy the sheet content and formatting
    for row in original_sheet.iter_rows():
        for cell in row:
            new_cell = new_sheet[cell.coordinate]
            new_cell.value = cell.value

            # Copy styles
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

            # Copy comments
            if cell.comment:
                new_cell.comment = copy(cell.comment)

    # Copy merged cells
    for merged_cell in original_sheet.merged_cells.ranges:
        new_sheet.merge_cells(str(merged_cell))

    # Copy sheet properties
    new_sheet.sheet_properties.tabColor = original_sheet.sheet_properties.tabColor

    new_file_path = os.path.join(output_directory, f'Chelsea Ng_Java-B {sheet_name}.xlsx')

    # Save the new workbook
    new_workbook.save(new_file_path)

print(
    "Sheets have been successfully exported to new Excel files in the 'report card' folder with formatting preserved.")
